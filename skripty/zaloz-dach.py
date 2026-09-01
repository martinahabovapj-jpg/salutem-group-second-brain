# -*- coding: utf-8 -*-
"""Zalozi prazdny master pro region DACH podle struktury ceskeho masteru.

PROC SE STRUKTURA KOPIRUJE, A NE VYMYSLI
Aplikace "Kdo mi to zafinancuje" i cely mesicni beh hledaji sloupce podle
nazvu z konfigurace. Kdyby se DACH sesit napsal rucne, lisil by se v jedne
mezere nebo diakritice a poznalo by se to az za mesic. Hlavicky se proto
ctou z ostreho ceskeho masteru a kopiruji znak po znaku.

CO SE KOPIRUJE
Nazvy listu, hlavicky sloupcu, sirky sloupcu a vysvetlivky pod tabulkou
u listu 7 a 8. DATA SE NEKOPIRUJI - databaze se zaklada prazdna. Ceske
subjekty do DACH sesitu nepatri a kopie by z nej udelala duplikat, ktery
by se za mesic rozesel s originalem.

LIST "0 Prehled" se nekopiruje, ale pise znovu - je v nem napsano cerne na
bilem, ze registrova kontrola u DACH NENI a proc.

POUZITI
    python zaloz-dach.py            # jen ukaze, co by vzniklo
    python zaloz-dach.py --zapis
"""

import argparse
import io
import json
import os
import shutil

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_CZ = os.path.join(HERE, "financovani-beh.config.json")
CONFIG_DACH = os.path.join(HERE, "financovani-beh-dach.config.json")
TEXTY = os.path.join(HERE, "zaloz-dach.texty.json")


def vypis(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def prehled(ws, T):
    """List 0 Prehled - pise se znovu, nekopiruje."""
    from openpyxl.styles import Font, Alignment
    ws["A1"] = T["nadpis"]
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = T["podnadpis"]
    r = 4
    for i, h in enumerate(T["hlavicka"]):
        c = ws.cell(row=r, column=i + 1)
        c.value = h
        c.font = Font(bold=True)
    r += 1
    for radek in T["radky"]:
        for i, v in enumerate(radek):
            ws.cell(row=r, column=i + 1).value = v
        r += 1
    r += 1
    ws.cell(row=r, column=1).value = T["proc_nadpis"]
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    for veta in T["proc"]:
        ws.cell(row=r, column=1).value = veta
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=False)
        r += 1
    r += 1
    ws.cell(row=r, column=1).value = T["pouziti_nadpis"]
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    for veta in T["pouziti"]:
        ws.cell(row=r, column=1).value = veta
        r += 1
    ws.column_dimensions["A"].width = 118
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60


def main():
    ap = argparse.ArgumentParser(description="Zalozi prazdny DACH master")
    ap.add_argument("--zapis", action="store_true")
    args = ap.parse_args()

    import openpyxl
    cfg_cz = json.load(io.open(CONFIG_CZ, encoding="utf-8"))
    cfg_dach = json.load(io.open(CONFIG_DACH, encoding="utf-8"))
    T = json.load(io.open(TEXTY, encoding="utf-8"))
    cil = cfg_dach["master"]

    if not os.path.isfile(cfg_cz["master"]):
        raise SystemExit("Cesky master neni videt - je disk O: pripojeny?\n  %s"
                         % cfg_cz["master"])
    if os.path.isfile(cil):
        raise SystemExit("DACH master uz existuje, nepreplacuji ho:\n  %s" % cil)

    zdroj = openpyxl.load_workbook(cfg_cz["master"])
    novy = openpyxl.Workbook()
    novy.remove(novy.active)

    vypis("Struktura se bere z: %s" % os.path.basename(cfg_cz["master"]))
    vypis("")
    for nazev in zdroj.sheetnames:
        ws_z = zdroj[nazev]
        ws_n = novy.create_sheet(nazev)
        if nazev.startswith("0 "):
            prehled(ws_n, T)
            vypis("  %-28s prehled napsan znovu" % nazev)
            continue
        hlavicka = [c.value for c in next(ws_z.iter_rows(min_row=1, max_row=1))]
        for i, h in enumerate(hlavicka, start=1):
            ws_n.cell(row=1, column=i).value = h
            zdroj_c = ws_z.cell(row=1, column=i)
            if zdroj_c.font and zdroj_c.font.bold:
                from openpyxl.styles import Font
                ws_n.cell(row=1, column=i).font = Font(bold=True)
        for pismeno, rozmer in ws_z.column_dimensions.items():
            if rozmer.width:
                ws_n.column_dimensions[pismeno].width = rozmer.width
        ws_n.freeze_panes = "A2"
        vypis("  %-28s %d sloupcu, 0 radku dat" % (nazev, len([h for h in hlavicka if h])))

    vypis("")
    vypis("Cil: %s" % cil)
    if not args.zapis:
        vypis("NANECISTO - nic nevzniklo. Pro zalozeni pridej --zapis")
        return
    slozka = os.path.dirname(cil)
    if not os.path.isdir(slozka):
        raise SystemExit("Cilova slozka neexistuje: %s" % slozka)
    novy.save(cil)
    vypis("Zalozeno.")


if __name__ == "__main__":
    main()
