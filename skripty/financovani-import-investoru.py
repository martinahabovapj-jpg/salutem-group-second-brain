# -*- coding: utf-8 -*-
"""Ze starych podkladu vytahne JEN JMENA A WEBY jako startovni seznam.

PROC EXISTUJE
Vedle masteru lezi 'Wealth management_FO.xlsx' - 90 firem z drivejsiho
zpusobu prace (cerven 2026). Jsou v nem i AUM, kontakty a strategie, ale
to jsou stara data a nepreklapi se.

Co v nich nezestarlo, jsou JMENA a WEBY. A prave jmeno je u family office
jedina vstupenka: vlastni NACE nemaji a v ARESu jsou k nerozeznani od bezne
s.r.o. Registr je podle oboru nenajde nikdy.

CO TENHLE SKRIPT DELA
  1) precte stary sesit a vezme z nej jmeno + web
  2) ke kazdemu dohleda ICO v ARESu (to je ziva informace, ne stara)
  3) odecte ty, ktere v masteru uz jsou - parovani pres domenu webu, ICO
     i nazev, protoze stejna firma je v obou seznamech psana jinak
  4) vyrobi seznam kandidatu k posouzeni

CO NEDELA
Nezapisuje do masteru. Neprenasi AUM, kontakty ani strategii ze stareho
sesitu. Ty musi vzniknout znovu z webu s doslovnou citaci - stejne jako
u kazdeho jineho subjektu v databazi.

DOHLEDANI ICO
Jmeno v podkladech je casto zkratka ('Odehnal & Partneri', 'Wealth Effect
Mgmt'). Postup je: cele jmeno -> presna shoda zacatku -> prvni vyznamove
slovo. Kdyz vyjde prave jeden zasah, bere se. Jinak zustane ICO prazdne
a rekne se to nahlas. Nikdy se nehada: radeji prazdne ICO nez cizi firma -
presne tahle chyba uz v tomhle projektu jednou nastala, kdyz se slovenske
subjekty ptaly ARESu a on odpovedel udaji o ceske firme.

POUZITI
    python financovani-import-investoru.py            # vyrobi seznam kandidatu
    python financovani-import-investoru.py --master X # proti jine kopii sesitu
"""

import argparse
import datetime as dt
import io
import json
import os
import re
import shutil
import socket
import ssl
import time
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG = os.path.join(HERE, "financovani-beh.config.json")
DNES = dt.date.today().isoformat()
URL_HLEDAT = "https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/vyhledat"

T = {}


def vypis(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def domena(url):
    """Porovnavaci klic webu. 'https://www.rb.cz/' i 'rb.cz' -> 'rb.cz'."""
    u = norm(url).lower()
    u = re.sub(r"^https?://", "", u)
    u = u.split()[0] if u else ""
    u = re.sub(r"^www\.", "", u)
    return u.split("/")[0].strip().strip(".")


def kontext():
    # soubor ma v nazvu pomlcky, takze se nacita pres loader, ne importem
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "financovani_beh", os.path.join(HERE, "financovani-beh.py"))
    modul = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modul)
    cfg = json.load(io.open(CONFIG, encoding="utf-8"))
    T.update(cfg["texty"])
    modul.T.update(cfg["texty"])
    return cfg, modul


# ---------------------------------------------------------------- zdroj

def nacti_podklady(cfg):
    import openpyxl
    i = cfg["import_investoru"]
    cesta = i["zdroj"]
    if not os.path.isfile(cesta):
        raise SystemExit(
            "Stary sesit s podklady nenalezen: %s\n"
            "Cesta se meni v financovani-beh.config.json, sekce import_investoru." % cesta)
    wb = openpyxl.load_workbook(cesta, read_only=True, data_only=True)
    ws = wb[i["list"]] if i.get("list") in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [norm(x) for x in rows[0]]
    mapa = i["sloupce"]
    out, videna = [], set()
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        firma = norm(d.get(mapa["firma"]))
        if not firma or firma.lower() in videna:
            continue
        videna.add(firma.lower())
        out.append({
            "firma": firma,
            "typ_stary": norm(d.get(mapa["typ"])),
            "web": norm(d.get(mapa["web"])).split()[0] if norm(d.get(mapa["web"])) else "",
        })
    return out, cesta


# ---------------------------------------------------------------- ICO

SLOVA_PRYC = re.compile(
    r"(?i)^(a\.?s\.?|s\.?r\.?o\.?|spol\.?|group|wealth|management|mgmt|partners?|"
    r"partneri|capital|invest\.?|investments?|bank|banka|fo|family|office|"
    r"private|sicav|holding|cz|sk|the|and)$")


def jadro_jmena(firma):
    """Prvni vyznamove slovo. 'Odehnal & Partneri' -> 'Odehnal'."""
    cist = re.sub(r"\(.*?\)", " ", firma).replace("&", " ").replace("-", " ")
    kusy = [k for k in re.split(r"[\s,./]+", cist) if len(k) > 2]
    for k in kusy:
        if not SLOVA_PRYC.match(k):
            return k
    return kusy[0] if kusy else ""


def ares_jmeno(jmeno, ctx, ua, pocet=30):
    telo = {"obchodniJmeno": jmeno, "start": 0, "pocet": pocet}
    req = urllib.request.Request(URL_HLEDAT, data=json.dumps(telo).encode("utf-8"),
                                 headers={"Content-Type": "application/json",
                                          "Accept": "application/json",
                                          "User-Agent": ua})
    try:
        d = json.load(urllib.request.urlopen(req, timeout=30, context=ctx))
    except Exception:
        return []
    return d.get("ekonomickeSubjekty", [])


def zjednodus(jmeno):
    """Nazev bez pravni formy, diakritiky a interpunkce. 'EMUN a.s.' -> 'emun'."""
    import unicodedata
    t = unicodedata.normalize("NFKD", norm(jmeno))
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    t = re.sub(r"\b(a\.?\s?s\.?|s\.?\s?r\.?\s?o\.?|spol\.?\s?s\s?r\.?\s?o\.?|"
               r"se|k\.?s\.?|v\.?o\.?s\.?|sicav|o\.?c\.?p\.?|z\.?s\.?|"
               r"investicni spolecnost|podfond)\b", " ", t)
    t = re.sub(r"[^a-z0-9]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def koren_domeny(web):
    """'https://www.emun.cz/' -> 'emun'. Domena se u firem casto rovna nazvu."""
    d = domena(web)
    if not d:
        return ""
    kusy = d.split(".")
    return kusy[0] if kusy else ""


def dohledej_ico(firma, web, ctx, ua):
    """Vraci (ico, jak, co_se_naslo). Nejednoznacnost = prazdne ICO.

    Poradi pokusu jde od nejsilnejsiho dokladu k nejslabsimu:
      1) jediny zasah na cele jmeno
      2) presna shoda zjednodusenych nazvu (bez pravni formy a diakritiky)
      3) shoda s korenem domeny webu - 'emun.cz' vs 'EMUN a.s.'
      4) jadro jmena, kdyz da prave jeden zasah
    Kdyz nic z toho neplati, ICO zustane prazdne. Radeji dira nez cizi firma.
    """
    kand = ares_jmeno(firma, ctx, ua)
    if len(kand) == 1:
        return norm(kand[0].get("ico")), "cele jmeno", norm(kand[0].get("obchodniJmeno"))

    cil = zjednodus(firma)
    shody = [z for z in kand if zjednodus(z.get("obchodniJmeno")) == cil]
    if len(shody) == 1:
        return norm(shody[0].get("ico")), "presna shoda nazvu", norm(shody[0].get("obchodniJmeno"))

    koren = koren_domeny(web)
    if koren and len(koren) > 2:
        pres_web = [z for z in kand
                    if zjednodus(z.get("obchodniJmeno")).replace(" ", "") == koren]
        if len(pres_web) == 1:
            return norm(pres_web[0].get("ico")), "shoda s domenou '%s'" % koren,                    norm(pres_web[0].get("obchodniJmeno"))
        if not kand:
            time.sleep(0.2)
            k3 = ares_jmeno(koren, ctx, ua)
            pres_web = [z for z in k3
                        if zjednodus(z.get("obchodniJmeno")).replace(" ", "") == koren]
            if len(pres_web) == 1:
                return norm(pres_web[0].get("ico")), "domena '%s'" % koren,                        norm(pres_web[0].get("obchodniJmeno"))

    jadro = jadro_jmena(firma)
    if jadro and jadro.lower() != firma.lower():
        time.sleep(0.2)
        k2 = ares_jmeno(jadro, ctx, ua)
        if len(k2) == 1:
            return norm(k2[0].get("ico")), "jadro '%s'" % jadro, norm(k2[0].get("obchodniJmeno"))
        shody2 = [z for z in k2 if zjednodus(z.get("obchodniJmeno")) == cil]
        if len(shody2) == 1:
            return norm(shody2[0].get("ico")), "jadro + shoda nazvu",                    norm(shody2[0].get("obchodniJmeno"))
        kand = kand or k2

    if not kand:
        return "", "ARES nenasel nic", ""
    return "", "%d zasahu, nejednoznacne" % len(kand),            "; ".join(norm(z.get("obchodniJmeno"))[:36] for z in kand[:3])


# ---------------------------------------------------------------- parovani

def rejstrik_masteru(sesit):
    """Tri klice na tutez firmu: domena webu, ICO, nazev. Stejna firma je
    v obou seznamech psana jinak ('Raiffeisenbank' vs 'SPM FINANCE ...')."""
    dom, ica, nazvy = {}, {}, {}
    for r, d in sesit.radky("subjekty"):
        zaznam = (r, norm(d.get("id")), norm(d.get("nazev")), norm(d.get("stav")))
        dm = domena(d.get("web"))
        if dm:
            dom.setdefault(dm, zaznam)
        for kus in re.findall(r"\d{8}", str(d.get("ico") or "")):
            ica.setdefault(kus, zaznam)
        nazvy.setdefault(norm(d.get("nazev")).lower(), zaznam)
    return dom, ica, nazvy


# ---------------------------------------------------------------- vystup

def zapis_kandidaty(vysledky, cfg, zdroj):
    d = os.path.join(HERE, cfg["import_investoru"]["k_posouzeni"])
    if os.path.isdir(d):
        shutil.rmtree(d)
    os.makedirs(d)

    novi = [v for v in vysledky if not v["v_masteru"]]
    znami = [v for v in vysledky if v["v_masteru"]]

    r = [T["import_nadpis"], "",
         T["import_uvod"].format(zdroj=os.path.basename(zdroj), celkem=len(vysledky),
                                 novych=len(novi), znamych=len(znami)),
         "", T["import_hlavicka"], "|---|---|---|---|---|"]
    for v in novi:
        r.append("| %s | %s | %s | %s | %s |" % (
            v["ico"] or "-", v["firma"], v["web"] or "-",
            v["typ_stary"] or "-", v["jak"]))
    r.append("")
    r.append(T["import_znami_nadpis"])
    r.append("")
    r.append(T["import_znami_hlavicka"])
    r.append("|---|---|---|")
    for v in znami:
        radek, sid, nazev, stav = v["v_masteru"]
        r.append("| #%s %s (%s) | %s | %s |" % (sid, nazev[:40], stav, v["firma"], v["parovano"]))
    io.open(os.path.join(d, "kandidati.md"), "w", encoding="utf-8").write("\n".join(r) + "\n")
    io.open(os.path.join(d, "_ZADANI.md"), "w", encoding="utf-8").write(T["import_zadani"])
    return len(novi), len(znami)


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(
        description="Startovni seznam investoru ze starych podkladu")
    ap.add_argument("--master", help="jina cesta k sesitu (test)")
    ap.add_argument("--limit", type=int, help="jen prvnich N firem (zkouska)")
    args = ap.parse_args()

    cfg, modul = kontext()
    if args.master:
        cfg["master"] = args.master
        vypis("POZOR: bezi proti jinemu sesitu nez podle konfigurace - %s" % args.master)
    sesit = modul.Sesit(cfg)

    firmy, zdroj = nacti_podklady(cfg)
    if args.limit:
        firmy = firmy[:args.limit]
    vypis("Ze starych podkladu nacteno firem: %d" % len(firmy))
    vypis("Prenasi se z nich JEN jmeno a web. AUM, kontakty a strategie ne -")
    vypis("ty musi vzniknout znovu z webu s citaci.")
    vypis("")

    dom, ica, nazvy = rejstrik_masteru(sesit)
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ctx = ssl.create_default_context()
    ua = cfg["sit"]["user_agent"]
    socket.setdefaulttimeout(cfg["sit"]["timeout_s"])

    vysledky = []
    for n, f in enumerate(firmy, 1):
        ico, jak, naslo = dohledej_ico(f["firma"], f["web"], ctx, ua)
        v_masteru, parovano = None, ""
        dm = domena(f["web"])
        if ico and ico in ica:
            v_masteru, parovano = ica[ico], T["parovano_ico"].format(ico=ico)
        elif dm and dm in dom:
            v_masteru, parovano = dom[dm], T["parovano_web"].format(dom=dm)
        elif f["firma"].lower() in nazvy:
            v_masteru, parovano = nazvy[f["firma"].lower()], T["parovano_nazev"]
        vysledky.append(dict(f, ico=ico, jak=jak, naslo=naslo,
                             v_masteru=v_masteru, parovano=parovano))
        vypis("  %3d/%d  %-34s ICO %-10s %s" % (n, len(firmy), f["firma"][:34],
                                                ico or "-", jak))
        time.sleep(0.15)

    novych, znamych = zapis_kandidaty(vysledky, cfg, zdroj)
    bez_ica = sum(1 for v in vysledky if not v["ico"])
    vypis("")
    vypis("=" * 62)
    vypis("  Celkem firem:                 %d" % len(vysledky))
    vypis("  Uz v masteru (nepridavat):    %d" % znamych)
    vypis("  Novych k posouzeni:           %d" % novych)
    vypis("  Z toho bez dohledaneho ICO:   %d" % bez_ica)
    vypis("=" * 62)
    vypis("Seznam je ve slozce %s" % cfg["import_investoru"]["k_posouzeni"])
    vypis("Nic nebylo zapsano do masteru - tenhle skript do nej nesaha.")


if __name__ == "__main__":
    main()
