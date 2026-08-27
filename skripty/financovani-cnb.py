# -*- coding: utf-8 -*-
"""Seznam investicnich fondu od CNB jako zdroj pro databazi financovani.

PROC EXISTUJE
ARES umi rict, ze fond vznikl. Neumi rict, co dela - u vsech fondu vraci
tyz NACE 64310. Web podfondu obvykle neexistuje: statut vydava az podfond
sam a nove podfondy o sobe tydny az mesice nezverejnuji nic. Posuzovat je
hned po zapisu do OR proto konci vetsinou u "nevim".

CNB vede mesicni Seznam investicnich fondu, ve kterem ma kazdy fond
KATEGORII podle investicni strategie - a jedna z nich je primo "uverovy".
To je presne to, na co se databaze pta, a je to udaj od regulatora,
ne marketingovy text z webu.

CO TENHLE SKRIPT DELA
  1) stahne seznam z cnb.cz (jeden xlsx, mesicni listy zpet do roku 2006)
  2) vezme nejnovejsi list
  3) sparuje ho s masterem pres ICO
  4) vypise, ktere fondy dane kategorie v databazi CHYBI
  5) na pozadani z nich vyrobi seznam kandidatu k posouzeni

CO NEDELA
Nezapisuje do masteru. Kategorie od CNB je silny doklad, ale porad to musi
projit posouzenim a schvalenim - stejne jako vsechno ostatni.

POZOR NA JEDNU VEC
Seznam je platny k poslednimu dni mesice. Fond zapsany do OR pozdeji v nem
jeste neni - proto se "neni v seznamu CNB" NESMI cist jako "neni fond".

POUZITI
    python financovani-cnb.py                          # prehled a mezery
    python financovani-cnb.py --kandidati              # + seznam k posouzeni
    python financovani-cnb.py --kategorie uverovy      # jen jedna kategorie
"""

import argparse
import io
import json
import os
import re
import shutil
import ssl
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG = os.path.join(HERE, "financovani-beh.config.json")

T = {}
SL = {}


def vypis(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def bez_diakritiky(s):
    import unicodedata
    t = unicodedata.normalize("NFKD", norm(s))
    return "".join(c for c in t if not unicodedata.combining(c)).lower()


def kontext():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "financovani_beh", os.path.join(HERE, "financovani-beh.py"))
    modul = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modul)
    cfg = json.load(io.open(CONFIG, encoding="utf-8"))
    T.update(cfg["texty"])
    SL.update(cfg["cnb"]["sloupce"])
    modul.T.update(cfg["texty"])
    return cfg, modul


def stahni(cfg):
    """Stahne seznam. Kdyz uz je stazeny z dneska, pouzije ho."""
    c = cfg["cnb"]
    cil = os.path.join(HERE, c["soubor"])
    if os.path.isfile(cil):
        import datetime as dt
        stari = dt.date.today().toordinal() - \
            dt.date.fromtimestamp(os.path.getmtime(cil)).toordinal()
        if stari < c["max_stari_dnu"]:
            vypis("Seznam CNB uz mame (stary %d dnu), nestahuji znovu." % stari)
            return cil
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ctx = ssl.create_default_context()
    vypis("Stahuji seznam investicnich fondu z CNB...")
    req = urllib.request.Request(c["url"], headers={"User-Agent": cfg["sit"]["user_agent"]})
    d = urllib.request.urlopen(req, timeout=120, context=ctx).read()
    with open(cil, "wb") as f:
        f.write(d)
    vypis("  ulozeno, %d kB" % (len(d) // 1024))
    return cil


def nacti(cesta):
    """Vraci (nazev listu, [{sloupec: hodnota}]). Bere nejnovejsi list."""
    import openpyxl
    wb = openpyxl.load_workbook(cesta, read_only=True, data_only=True)
    list_n = wb.sheetnames[0]          # CNB radi nejnovejsi mesic jako prvni
    ws = wb[list_n]
    rows = list(ws.iter_rows(values_only=True))
    # hlavicka neni na prvnim radku - najde se podle sloupce RIAD_CODE
    hi = next(i for i, r in enumerate(rows) if r and norm(r[0]) == SL["hlavicka_prvni"])
    hdr = [norm(x) for x in rows[hi]]
    data = [dict(zip(hdr, r)) for r in rows[hi + 1:] if r and r[2]]
    platnost = norm(rows[1][0]) if len(rows) > 1 else ""
    return list_n, platnost, data


def ica_masteru(sesit):
    out = {}
    for r, d in sesit.radky("subjekty"):
        for kus in re.findall(r"\d{8}", str(d.get("ico") or "")):
            out[kus] = (norm(d.get("id")), norm(d.get("nazev")), norm(d.get("stav")))
    return out



# ------------------------------------------------------- investorska strana
# Nemovitostni fondy nejsou veritele - jsou to alokatori kapitalu do
# nemovitosti, tedy list 6. Je jich ale 190 a nekolik jich casto patri
# tomuz spravci: WOOD & Company ma sedm fondu, RSJ dvanact ve trech vetvich.
# Oslovuje se spravce, ne kazdy podfond zvlast, takze se sem chodi
# PO SKUPINACH. Jinak by fronta mela sedm radku pro tutez firmu a
# schvalovatel by se ji naucil preskakovat.


def skupina(nazev, vyhodit, pocet):
    """Klic rodiny fondu. 'WOOD & Company Retail podfond' -> 'WOOD COMPANY'."""
    n = norm(nazev)
    for slovo in vyhodit:
        n = re.sub(r"(?i)\b%s\b" % re.escape(slovo), " ", n)
    n = re.sub(r"[^0-9A-Za-z\u00c0-\u017e ]", " ", n)
    kusy = []
    for x in n.split():
        # zdvojene slovo v nazvu ("Velvet SICAV podfond Velvet Hotel") by
        # z klice udelalo "VELVET VELVET" - beru kazde slovo jen jednou
        if len(x) > 1 and (not kusy or x.upper() != kusy[-1].upper()):
            kusy.append(x)
    return " ".join(kusy[:pocet]).upper() if kusy else "?"


def po_skupinach(data, znama, cfg, kategorie):
    """Vraci [{skupina, fondy:[...]}] serazene podle poctu fondu."""
    vyhodit = cfg["cnb"]["investor"]["vyhodit_ze_jmena"]
    pocet = cfg["cnb"]["investor"].get("slov_ve_skupine", 1)
    kb = bez_diakritiky(kategorie)
    chybi = [d for d in data
             if bez_diakritiky(d.get(SL["kategorie"])) == kb
             and norm(d.get(SL["ico"])) not in znama]
    mapa = {}
    for d in chybi:
        mapa.setdefault(skupina(d.get(SL["nazev"]), vyhodit, pocet), []).append(d)
    out = []
    for k in mapa:
        fondy = sorted(mapa[k], key=lambda x: norm(x.get(SL["nazev"])))
        # reprezentant: radeji samostatny fond nez podfond - u podfondu je
        # protistranou stejne az spravce, ktereho v tomhle seznamu nemame
        sam = [f for f in fondy if norm(f.get(SL["podfond"])).upper() == "NE"]
        rep = (sam or fondy)[0]
        out.append({"skupina": k, "fondy": fondy, "rep": rep,
                    "samostatnych": len(sam)})
    out.sort(key=lambda s: (-len(s["fondy"]), s["skupina"]))
    return out


def zapis_skupiny(skupiny, cfg, list_n, platnost, kategorie):
    d_out = os.path.join(HERE, cfg["cnb"]["investor"]["k_posouzeni"])
    if os.path.isdir(d_out):
        shutil.rmtree(d_out)
    os.makedirs(d_out)
    fondu = sum(len(s["fondy"]) for s in skupiny)
    r = [T["cnb_inv_nadpis"], "",
         T["cnb_inv_uvod"].format(list=list_n, platnost=platnost,
                                  kategorie=kategorie, fondu=fondu,
                                  skupin=len(skupiny)),
         "", T["cnb_inv_hlavicka"], "|---|---|---|---|---|"]
    for s in skupiny:
        rep = s["rep"]
        ica = ", ".join(norm(f.get(SL["ico"])) for f in s["fondy"][:6])
        if len(s["fondy"]) > 6:
            ica += ", ..."
        r.append("| %s | %d | %s | %s | %s |" % (
            s["skupina"], len(s["fondy"]), norm(rep.get(SL["nazev"]))[:70],
            "%s, %s" % (norm(rep.get(SL["mesto"])), norm(rep.get(SL["ulice"]))),
            ica))
    io.open(os.path.join(d_out, "kandidati.md"), "w", encoding="utf-8").write(
        "\n".join(r) + "\n")
    io.open(os.path.join(d_out, "_ZADANI.md"), "w", encoding="utf-8").write(
        T["cnb_inv_zadani"])
    return len(skupiny), fondu


def main():
    ap = argparse.ArgumentParser(description="Seznam investicnich fondu CNB proti databazi")
    ap.add_argument("--kategorie", help="jen tyhle kategorie, oddelene carkou "
                                        "(bez diakritiky, napr. uverovy,nemovitostni)")
    ap.add_argument("--kandidati", action="store_true",
                    help="vyrobi seznam chybejicich fondu k posouzeni")
    ap.add_argument("--master", help="jina cesta k sesitu (test)")
    ap.add_argument("--investor", action="store_true",
                    help="investorska strana (list 6): nemovitostni fondy PO SKUPINACH")
    args = ap.parse_args()

    cfg, modul = kontext()
    if args.master:
        cfg["master"] = args.master
    sesit = modul.Sesit(cfg)

    cesta = stahni(cfg)
    list_n, platnost, data = nacti(cesta)
    vypis("Seznam CNB: list '%s', %s, fondu: %d" % (list_n, platnost, len(data)))

    if args.investor:
        kat = cfg["cnb"]["investor"]["kategorie"]
        znama = ica_masteru(sesit)
        skupiny = po_skupinach(data, znama, cfg, kat)
        vypis("")
        vypis("=" * 70)
        vypis("  " + T["nadpis_cnb_investor"])
        vypis("=" * 70)
        for s in skupiny[:25]:
            vypis("   %-28s %d fondu   %s" % (s["skupina"], len(s["fondy"]),
                                              norm(s["rep"].get(SL["nazev"]))[:44]))
        if len(skupiny) > 25:
            vypis("   ... a dalsich %d skupin s malym poctem fondu" % (len(skupiny) - 25))
        n, f = zapis_skupiny(skupiny, cfg, list_n, platnost, kat)
        vypis("=" * 70)
        vypis("Kategorie '%s': %d fondu, ktere nemame, tvori %d skupin." % (kat, f, n))
        vypis("K posouzeni pripraveno %d skupin (slozka %s)"
              % (n, cfg["cnb"]["investor"]["k_posouzeni"]))
        vypis("Oslovuje se sprava skupiny, ne kazdy podfond zvlast.")
        return

    chci = cfg["cnb"]["kategorie_zajmu"]
    if args.kategorie:
        chci = [x.strip() for x in args.kategorie.split(",") if x.strip()]
    chci_bd = [bez_diakritiky(x) for x in chci]

    znama = ica_masteru(sesit)
    vypis("")
    vypis("=" * 70)
    for kat in chci:
        kb = bez_diakritiky(kat)
        skupina = [d for d in data if bez_diakritiky(d.get(SL["kategorie"])) == kb]
        mame = [d for d in skupina if norm(d.get(SL["ico"])) in znama]
        nemame = [d for d in skupina if norm(d.get(SL["ico"])) not in znama]
        vypis("  KATEGORIE '%s': v CR %d, v databazi %d, CHYBI %d"
              % (kat, len(skupina), len(mame), len(nemame)))
        for d in sorted(nemame, key=lambda x: norm(x.get(SL["nazev"]))):
            vypis("     %-9s %-56s podfond=%s"
                  % (norm(d.get(SL["ico"])), norm(d.get(SL["nazev"]))[:56], d.get(SL["podfond"])))
    vypis("=" * 70)

    if not args.kandidati:
        vypis("Nic nezapsano. Seznam k posouzeni vyrobis prepinacem --kandidati")
        return

    d_out = os.path.join(HERE, cfg["cnb"]["k_posouzeni"])
    if os.path.isdir(d_out):
        shutil.rmtree(d_out)
    os.makedirs(d_out)
    radky = [T["cnb_nadpis"], "",
             T["cnb_uvod"].format(list=list_n, platnost=platnost, celkem=len(data)),
             "", T["cnb_hlavicka"], "|---|---|---|---|---|"]
    pocet = 0
    for kat, kb in zip(chci, chci_bd):
        for d in sorted(data, key=lambda x: norm(x.get(SL["nazev"]))):
            if bez_diakritiky(d.get(SL["kategorie"])) != kb:
                continue
            if norm(d.get(SL["ico"])) in znama:
                continue
            radky.append("| %s | %s | %s | %s | %s |" % (
                norm(d.get(SL["ico"])), norm(d.get(SL["nazev"])), kat,
                d.get(SL["podfond"]), "%s, %s" % (norm(d.get(SL["mesto"])), norm(d.get(SL["ulice"])))))
            pocet += 1
    io.open(os.path.join(d_out, "kandidati.md"), "w", encoding="utf-8").write(
        "\n".join(radky) + "\n")
    io.open(os.path.join(d_out, "_ZADANI.md"), "w", encoding="utf-8").write(T["cnb_zadani"])
    vypis("K posouzeni pripraveno: %d (slozka %s)" % (pocet, cfg["cnb"]["k_posouzeni"]))


if __name__ == "__main__":
    main()
