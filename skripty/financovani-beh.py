# -*- coding: utf-8 -*-
"""
financovani-beh.py - mesicni beh nad Master databazi poskytovatelu financovani.

Co to dela
----------
Projde subjekty v master sesitu, zjisti, co se od minule zmenilo, a kazdou
zmenu zaradi do jednoho ze tri pruhu:

    A  projde sama (registrove a sitove veci, ktere nema smysl davat cloveku)
    B  jde do listu "Navrhy zmen" ke schvaleni
    C  jen do logu

Routuje se podle toho, KTERE POLE se meni a ODKUD informace prisla -
nikdy podle obsahu zmeny. Pravidla jsou v financovani-beh.config.json,
sekce "pruhy". Kdyz neco chodi do spatneho pruhu, prehodi se to tam, ne tady.

Faze behu
---------
    0  nacteni sesitu
    1  registry: ARES + ISIR podle ICO          (bez modelu, zdroj celeho pruhu A)
    2  dostupnost: HTTP status webu, DNS mailove domeny   (bez modelu)
    3  otisk stranek: stahne, ocisti, sha256, porovna s minulym behem (bez modelu)
    4  cteni: jen zmenene stranky jdou modelu   -> viz --navrhy nize
    5  routovani a zapis

Pouziti
-------
    python financovani-beh.py                     # nanecisto, nic nezapise
    python financovani-beh.py --zapis             # ostry beh
    python financovani-beh.py --jen-registry      # jen faze 1 (nejlevnejsi a nejspolehlivejsi)
    python financovani-beh.py --limit 10          # jen prvnich 10 subjektu (test)
    python financovani-beh.py --navrhy navrhy.json --zapis   # prijme navrhy z faze 4
    python financovani-beh.py --vrat 2026-09-01   # vrati automaticke zmeny z daneho behu

Faze 4 (cteni modelem) neni v tomhle skriptu zamerne. Skript pripravi do slozky
"k-precteni" rozdil kazde zmenene stranky a zadani; model (Claude Code, skill
financovani-mesicni-beh) z toho vyrobi navrhy.json a ten se vrati sem prepinacem
--navrhy. Duvod: deterministicke kroky nepatri modelu a model nesmi psat do sesitu.

Stav mezi behy je v financovani-beh.stav.json. Bez nej neexistuje pojem
"zmenilo se od minule" a beh by pokazde navrhoval vsechno znovu.
"""

import argparse
import datetime as dt
import difflib
import hashlib
import json
import os
import re
import shutil
import socket
import ssl
import sys
import time
import urllib.error
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG = os.path.join(HERE, "financovani-beh.config.json")
STATE = os.path.join(HERE, "financovani-beh.stav.json")

DNES = dt.date.today().isoformat()

# Vsechny retezce s diakritikou jsou v konfiguraci, sekce "texty".
# Duvod: tenhle soubor zustava ciste ASCII - jinak se pri kazdem prusunu
# pres editor nebo PowerShell rozsype kodovani a do sesitu se zapise pasklika.
T = {}


# ---------------------------------------------------------------- pomocne

def load_json(path, default):
    if not os.path.isfile(path):
        return default
    # utf-8-sig: Windows editory pisou BOM, na kterem by json.load spadl
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def vypis(text):
    # konzole na Windows umi shodit non-ASCII, tak radeji opatrne
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", "replace").decode("ascii"))


def prazdne(v):
    return v is None or str(v).strip() in ("", "-", chr(8212))


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


# ---------------------------------------------------------------- faze 0

class Sesit(object):
    """Master sesit. Sloupce se hledaji podle nazvu z konfigurace, ne podle poradi."""

    def __init__(self, cfg):
        import openpyxl
        self.cfg = cfg
        self.cesta = cfg["master"]
        if not os.path.isfile(self.cesta):
            # Rozlisit "disk je pryc" od "cesta je spatne". Sitovy disk se
            # odpojuje sam od sebe a poslat kolegu opravovat konfiguraci,
            # kdyz staci pockat nebo obnovit pripojeni, je spatna rada.
            koren = os.path.splitdrive(self.cesta)[0]
            if koren and not os.path.isdir(koren + os.sep):
                raise SystemExit(
                    "Disk %s neni pripojeny, takze na master sesit nevidime.\n"
                    "Konfigurace je v poradku - nic v ni nemen.\n"
                    "Otevri Pruzkumnik souboru, klikni na %s a spust beh znovu.\n"
                    "Kdyz se disk neobjevi, je problem v siti nebo na serveru."
                    % (koren, koren))
            raise SystemExit("Master sesit nenalezen: %s\n"
                             "Zkontroluj, jestli soubor nekdo neprejmenoval nebo "
                             "neprestehoval. Cesta se meni v financovani-beh.config.json."
                             % self.cesta)
        self.wb = openpyxl.load_workbook(self.cesta)
        self.listy = cfg["listy"]
        self._hlavicky = {}

    def ws(self, klic):
        nazev = self.listy[klic]
        if nazev not in self.wb.sheetnames:
            raise SystemExit("V sesitu chybi list '%s'. Sloupce a listy se nepremenovavaji "
                             "- viz pravidlo v konfiguraci." % nazev)
        return self.wb[nazev]

    def hlavicka(self, klic):
        """{nazev sloupce: index od 1}"""
        if klic in self._hlavicky:
            return self._hlavicky[klic]
        ws = self.ws(klic)
        h = {}
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
            if c.value is not None:
                h[norm(c.value)] = i
        self._hlavicky[klic] = h
        return h

    def sl(self, klic, pole):
        """Index sloupce podle logickeho nazvu pole z konfigurace.

        Vraci None i tehdy, kdyz pole v konfiguraci daneho listu vubec neni -
        list 3 nema sloupec Overeno a ptat se na nej je legitimni.
        """
        nazev = self.cfg["sloupce"][klic].get(pole)
        if not nazev:
            return None
        h = self.hlavicka(klic)
        if nazev not in h:
            return None
        return h[nazev]

    def radky(self, klic):
        """[(cislo radku, {pole: hodnota})] pro vsechny neprazdne radky."""
        ws = self.ws(klic)
        mapa = self.cfg["sloupce"][klic]
        idx = {}
        for pole in mapa:
            i = self.sl(klic, pole)
            if i:
                idx[pole] = i
        out = []
        for r in range(2, ws.max_row + 1):
            data = {}
            for pole, i in idx.items():
                data[pole] = ws.cell(row=r, column=i).value
            if all(prazdne(v) for v in data.values()):
                continue
            out.append((r, data))
        return out

    def uloz(self, zaloha_dir):
        if zaloha_dir:
            if not os.path.isdir(zaloha_dir):
                os.makedirs(zaloha_dir)
            zal = os.path.join(zaloha_dir, "%s__pred-behem-%s.xlsx" % (
                os.path.splitext(os.path.basename(self.cesta))[0], DNES))
            if not os.path.isfile(zal):
                shutil.copy2(self.cesta, zal)
        try:
            self.wb.save(self.cesta)
        except PermissionError:
            raise SystemExit(
                "Sesit je zamceny - nekdo ho ma otevreny v Excelu.\n"
                "Zadna zmena nebyla zapsana. Zavri Excel a spust beh znovu\n"
                "(nebo nech beh jet v noci, kdy je soubor volny).")


# ---------------------------------------------------------------- sit

class Sit(object):
    def __init__(self, cfg):
        s = cfg["sit"]
        self.timeout = s["timeout_s"]
        self.pauza = s["pauza_mezi_dotazy_s"]
        self.ua = s["user_agent"]
        self.ep = cfg["endpointy"]
        socket.setdefaulttimeout(self.timeout)
        # statistics.sk posila retezec certifikatu, ktery Windows sam neoveri;
        # certifi ma vlastni seznam autorit a projde. Bez nej to spadne na
        # CERTIFICATE_VERIFY_FAILED a vypadalo by to jako "SK registr nejde".
        try:
            import certifi
            self.ctx = ssl.create_default_context(cafile=certifi.where())
        except ImportError:
            self.ctx = ssl.create_default_context()

    def _get(self, url, data=None, hlavicky=None):
        h = {"User-Agent": self.ua}
        if hlavicky:
            h.update(hlavicky)
        req = urllib.request.Request(url, data=data, headers=h)
        r = urllib.request.urlopen(req, timeout=self.timeout, context=self.ctx)
        return r.getcode(), r.read()

    # ---- faze 1c: RPO (Slovensko)
    def rpo(self, ico):
        """Slovensky protejsek ARESu. Vraci stejny tvar, aby se s tim dalo
        pracovat stejne - jen bez priznaku insolvence, ten SK registr nenese."""
        time.sleep(self.pauza)
        url = self.ep["rpo_sk_hledani"].format(ico=ico)
        try:
            kod, telo = self._get(url)
        except urllib.error.HTTPError as e:
            return {"ok": False, "chyba": "HTTP %s" % e.code}
        except Exception as e:
            return {"ok": False, "chyba": type(e).__name__}
        d = json.loads(telo.decode("utf-8"))
        vysledky = d.get("results") or []
        if not vysledky:
            return {"ok": False, "chyba": "nenalezeno"}
        r = vysledky[0]
        jmena = r.get("fullNames") or [{}]
        adresy = r.get("addresses") or [{}]
        a = adresy[0]
        obec = (a.get("municipality") or {}).get("value", "")
        return {
            "ok": True,
            "nazev": norm(jmena[0].get("value")),
            "sidlo": norm(" ".join(str(x) for x in [a.get("street"),
                                                    a.get("buildingNumber"),
                                                    obec] if x)),
            "zanik": r.get("termination"),
            "vr": None,
            "ir": None,   # slovenska insolvence je jiny registr, zatim nepokryta
        }

    # ---- faze 1a: ARES
    def ares(self, ico):
        """{'ok', 'nazev', 'sidlo', 'zanik', 'ir', 'vr'} nebo {'chyba': ...}"""
        time.sleep(self.pauza)
        url = self.ep["ares"].format(ico=ico)
        try:
            kod, telo = self._get(url)
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return {"ok": False, "chyba": "nenalezeno"}
            return {"ok": False, "chyba": "HTTP %s" % e.code}
        except Exception as e:
            return {"ok": False, "chyba": type(e).__name__}
        d = json.loads(telo.decode("utf-8"))
        sidlo = d.get("sidlo") or {}
        reg = d.get("seznamRegistraci") or {}
        return {
            "ok": True,
            "nazev": norm(d.get("obchodniJmeno")),
            "sidlo": norm(sidlo.get("textovaAdresa")
                          or " ".join(str(x) for x in [sidlo.get("nazevUlice"),
                                                       sidlo.get("cisloDomovni"),
                                                       sidlo.get("nazevObce")] if x)),
            "zanik": d.get("datumZaniku"),
            "vr": reg.get("stavZdrojeVr"),
            "ir": reg.get("stavZdrojeIr"),
        }

    # ---- faze 1b: ISIR
    def isir(self, ico):
        """{'v_insolvenci': bool, 'sz', 'soud', 'stav', 'url', 'datum'}"""
        time.sleep(self.pauza)
        telo = (
            '<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" '
            'xmlns:typ="http://isirws.cca.cz/types/"><soapenv:Header/><soapenv:Body>'
            "<typ:getIsirWsCuzkDataRequest><ic>%s</ic>"
            "<maxPocetVysledku>20</maxPocetVysledku>"
            "<filtrAktualniRizeni>T</filtrAktualniRizeni>"
            "</typ:getIsirWsCuzkDataRequest></soapenv:Body></soapenv:Envelope>" % ico
        ).encode("utf-8")
        try:
            kod, odp = self._get(self.ep["isir"], data=telo,
                                 hlavicky={"Content-Type": "text/xml;charset=UTF-8",
                                           "SOAPAction": ""})
        except Exception as e:
            return {"chyba": type(e).__name__}
        t = odp.decode("utf-8", "replace")

        def tag(n):
            m = re.search(r"<%s>(.*?)</%s>" % (n, n), t, re.S)
            return norm(m.group(1)) if m else ""

        if "<data>" not in t:
            return {"v_insolvenci": False}
        sz = "%s %s/%s" % (tag("druhVec"), tag("bcVec"), tag("rocnik"))
        return {
            "v_insolvenci": True,
            "sz": sz.strip(),
            "soud": tag("nazevOrganizace"),
            "stav": tag("druhStavKonkursu"),
            "url": tag("urlDetailRizeni"),
            "datum": tag("datumPmZahajeniUpadku"),
        }

    # ---- faze 2: dostupnost
    def http_stav(self, url):
        time.sleep(self.pauza)
        try:
            kod, telo = self._get(url)
            return kod, telo
        except urllib.error.HTTPError as e:
            return e.code, b""
        except Exception:
            return 0, b""

    def dns_ok(self, domena):
        try:
            socket.getaddrinfo(domena, None)
            return True
        except Exception:
            return False


# ---------------------------------------------------------------- faze 3

SKRIPTY = re.compile(r"<(script|style|noscript)[^>]*>.*?</\1>", re.S | re.I)
KOMENTARE = re.compile(r"<!--.*?-->", re.S)
TAGY = re.compile(r"<[^>]+>")
DATUMY = re.compile(r"\b\d{1,2}\.\s?\d{1,2}\.\s?\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b")
CASY = re.compile(r"\b\d{1,2}:\d{2}(:\d{2})?\b")
TOKENY = re.compile(r"\b[a-f0-9]{16,}\b", re.I)
COOKIE = re.compile(r"(?i)(cookie|souhlas s pouzivanim|nastaveni soukromi)[^.]{0,200}\.")


def ocisti(html_bytes, vzory=()):
    """Text stranky bez balastu, ktery se meni sam od sebe.

    Vyhazuje skripty, styly, komentare, datumy, casy, hashe a cookie listy,
    plus vzory z konfigurace (sekce otisk_ignoruj) - tam patri veci jako
    pocitaci captcha, ktera je pri kazdem nacteni jina.

    Cisla se NEvyhazuji - zmena ticketu nebo LTV je presne to, co hledame.
    """
    try:
        t = html_bytes.decode("utf-8")
    except UnicodeDecodeError:
        t = html_bytes.decode("cp1250", "replace")
    t = SKRIPTY.sub(" ", t)
    t = KOMENTARE.sub(" ", t)
    t = TAGY.sub("\n", t)
    t = COOKIE.sub(" ", t)
    t = DATUMY.sub(" ", t)
    t = CASY.sub(" ", t)
    t = TOKENY.sub(" ", t)
    for v in vzory:
        t = re.sub(v, " ", t)
    radky = [re.sub(r"[ \t]+", " ", r).strip() for r in t.splitlines()]
    radky = [r for r in radky if r]
    return "\n".join(radky)


def otisk(text):
    """Otisk se pocita ze SERAZENYCH unikatnich radku, ne z textu v poradi.

    Nekolik webu micha poradi poli ve formularich pri kazdem nacteni jako
    ochranu proti robotum. Bez tohohle by takova stranka hlasila zmenu kazdy
    mesic, model by u ni pokazde rekl "nic" a fronta by se zaplnila sumem.
    """
    klic = "\n".join(sorted(set(text.splitlines())))
    return hashlib.sha256(klic.encode("utf-8")).hexdigest()[:16]


# ---------------------------------------------------------------- navrhy

def navrh(id_subjektu, subjekt, druh, co, bylo, navrzeno, zdroj, citace="", jistota=""):
    return {
        "id": id_subjektu,
        "subjekt": subjekt,
        "druh": druh,          # klic do cfg["pruhy"] - podle nej se routuje
        "co": co,              # lidsky nazev pole
        "bylo": bylo,
        "navrzeno": navrzeno,
        "zdroj": zdroj,
        "citace": citace,
        "jistota": jistota,
    }


def routuj(n, cfg):
    """Pruh se urcuje podle DRUHU zmeny (jake pole, odkud), nikdy podle obsahu."""
    return cfg["pruhy"].get(n["druh"], "B")


def pod_prahem(n, cfg):
    """True = zmena je prilis mala na to, aby obtezovala schvalovatele (-> pruh C)."""
    p = cfg["prahy"]
    if n["druh"] == "ticket":
        a, b = cislo(n["bylo"]), cislo(n["navrzeno"])
        if a and b:
            return abs(b - a) / float(a) * 100 < p["ticket_procent"]
    if n["druh"] == "ltv":
        a, b = cislo(n["bylo"]), cislo(n["navrzeno"])
        if a is not None and b is not None:
            return abs(b - a) < p["ltv_procentni_body"]
    return False


def cislo(v):
    if v is None:
        return None
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(v).replace(" ", ""))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return None


# ---------------------------------------------------------------- faze 1+2+3

def projdi(sesit, sit, stav, cfg, limit=None, jen_registry=False):
    """Vrati (navrhy, zmenene_stranky). Nic nezapisuje."""
    scfg = cfg["sloupce"]["subjekty"]
    stavy = cfg["stavy"]
    navrhy = []
    zmenene = []
    subjekty = sesit.radky("subjekty")
    if limit:
        subjekty = subjekty[:limit]

    # vyrazene subjekty se v registrech nekontroluji kazdy mesic - jsou vyrazene
    zive = [(r, d) for r, d in subjekty
            if norm(d.get("stav")) != stavy["vyrazen"]]

    vypis("Faze 0: sesit nacten, subjektu %d, z toho nevyrazenych %d"
          % (len(subjekty), len(zive)))

    st_subj = stav.setdefault("subjekty", {})
    preskoceno_ico = []
    preskocena_zeme = []
    bez_zeme = []
    podle_zeme = {k: v for k, v in cfg["registry_podle_zeme"].items()
                  if not k.startswith("_")}

    # ---- faze 1: registry
    vypis("Faze 1: registry...")
    for i, (radek, d) in enumerate(zive, start=1):
        sid = str(d.get("id"))
        s = st_subj.setdefault(sid, {})
        ico = norm(d.get("ico"))
        nazev = norm(d.get("nazev"))
        zeme = norm(d.get("zeme")).upper()

        # ARES je cesky registr. Na slovenske ICO vraci 404 - a to NENI doklad
        # o zaniku firmy, je to jen dukaz, ze se ptame spatneho registru.
        registr = podle_zeme.get(zeme)
        if registr is None:
            # Dva ruzne problemy, ktere se nesmi slit do jedne hlasky:
            # "zeme nema napojeny registr" je vlastnost sveta a nic s tim
            # nenadelame; "zeme neni vyplnena" je dira v datech a spravit
            # se da. Kdyz se hlasi stejne, ta druha se ztrati mezi prvni.
            if not zeme:
                bez_zeme.append((sid, nazev))
            else:
                preskocena_zeme.append((sid, nazev, zeme))
            continue
        if not re.match(r"^\d{6,8}$", ico):
            preskoceno_ico.append((sid, nazev, ico))
            continue

        if registr == "rpo":
            a = sit.rpo(ico)
            odkaz = cfg["endpointy"]["rpo_sk_hledani"].format(ico=ico)
            chybi_text = T["rpo_nesedi"]
        else:
            a = sit.ares(ico)
            odkaz = cfg["endpointy"]["ares"].format(ico=ico)
            chybi_text = T["ico_nesedi"]

        if not a.get("ok"):
            if a.get("chyba") == "nenalezeno":
                # 404 neznamena zaniklou firmu, znamena ICO, ktere nesedi -> clovek
                navrhy.append(navrh(sid, nazev, "ico_nesedi", T["pole_ico"],
                                    ico, chybi_text, odkaz))
            s["ares_chyba"] = a.get("chyba")
            continue
        s.pop("ares_chyba", None)

        # zanik / likvidace
        if a.get("zanik") or a.get("vr") == "ZANIKLY":
            if norm(d.get("stav")) != stavy["zanikly"]:
                navrhy.append(navrh(sid, nazev, "zanik", "Stav",
                                    norm(d.get("stav")), stavy["zanikly"],
                                    odkaz,
                                    citace="datumZaniku=%s" % a.get("zanik")))
        elif a["nazev"].lower().endswith("v likvidaci"):
            if norm(d.get("stav")) != stavy["likvidace"]:
                navrhy.append(navrh(sid, nazev, "likvidace", "Stav",
                                    norm(d.get("stav")), stavy["likvidace"],
                                    odkaz,
                                    citace=a["nazev"]))

        # zmena nazvu a sidla se hlasi jen proti minulemu behu, ne proti nasemu
        # oznaceni v sesitu - nase nazvy nesou dodatky ("Podfond Loan") a
        # porovnavat je s ARESem by vyrobilo stovky falesnych nalezu.
        if s.get("ares_nazev") and s["ares_nazev"] != a["nazev"]:
            navrhy.append(navrh(sid, nazev, "zmena_nazvu", T["pole_nazev_rejstrik"],
                                s["ares_nazev"], a["nazev"],
                                odkaz))
        if s.get("ares_sidlo") and a["sidlo"] and s["ares_sidlo"] != a["sidlo"]:
            navrhy.append(navrh(sid, nazev, "zmena_sidla", T["pole_sidlo"],
                                s["ares_sidlo"], a["sidlo"],
                                odkaz))
        s["ares_nazev"] = a["nazev"]
        s["ares_sidlo"] = a["sidlo"]
        s["ares_ir"] = a.get("ir")

        # insolvence: pojistka - musi sedet v ARESu i v ISIRu
        ares_rika_ir = (a.get("ir") == "AKTIVNI")
        if ares_rika_ir or not cfg["pojistky"]["insolvence_vyzaduje_ares_i_isir"]:
            iz = sit.isir(ico)
            if iz.get("v_insolvenci"):
                if norm(d.get("stav")) != stavy["insolvence"]:
                    navrhy.append(navrh(
                        sid, nazev, "insolvence", "Stav",
                        norm(d.get("stav")), stavy["insolvence"],
                        iz.get("url") or cfg["endpointy"]["isir"],
                        citace="%s, %s, %s, uvod %s" % (iz.get("sz"), iz.get("soud"),
                                                        iz.get("stav"), iz.get("datum"))))
            s["isir"] = iz.get("v_insolvenci", False)
        s["posledni_registr"] = DNES

        if i % 20 == 0:
            vypis("  ... %d/%d" % (i, len(zive)))

    # Mlceni by cetlo jako "zkontrolovano". Co beh nepokryl, musi rict nahlas.
    if preskoceno_ico:
        vypis("  ICO nema tvar, na ktery jde registr zeptat, u %d subjektu - "
              "patri cloveku:" % len(preskoceno_ico))
        for sid, nazev, ico in preskoceno_ico:
            vypis("    #%s %s -> '%s'" % (sid, nazev, ico))
    if preskocena_zeme:
        vypis("  Zeme bez napojeneho registru u %d subjektu - v registrech se "
              "NEKONTROLUJI vubec:" % len(preskocena_zeme))
        for sid, nazev, zeme in preskocena_zeme:
            vypis("    #%s %s (%s)" % (sid, nazev, zeme))
    if bez_zeme:
        vypis("  ZEME NENI VYPLNENA u %d subjektu - neni podle ceho vybrat "
              "registr. Tohle jde spravit, staci doplnit sloupec Zeme:"
              % len(bez_zeme))
        for sid, nazev in bez_zeme:
            vypis("    #%s %s" % (sid, nazev))

    if jen_registry:
        return navrhy, zmenene

    # ---- faze 2 + 3: dostupnost a otisk
    vypis("Faze 2+3: dostupnost webu a otisk stranek...")
    nepusteni = []
    vzory_sumu = cfg.get("otisk_ignoruj", {}).get("vzory", [])
    nesledovane = {k: v for k, v in cfg.get("otisk_ignoruj", {})
                   .get("subjekty", {}).items() if not k.startswith("_")}
    stranky_dir = os.path.join(HERE, cfg["stranky"])
    if not os.path.isdir(stranky_dir):
        os.makedirs(stranky_dir)

    for radek, d in zive:
        sid = str(d.get("id"))
        s = st_subj.setdefault(sid, {})
        nazev = norm(d.get("nazev"))
        web = norm(d.get("web"))
        if not web.startswith("http"):
            continue

        kod, telo = sit.http_stav(web)
        # 401/403/429 neznamena "web nefunguje", znamena "nepustili nas dovnitr".
        # Je to NEVIM, ne NE - a splacnout to dohromady by vyrobilo poznamku
        # "web nedostupny" u firmy, ktera ma web v naprostem poradku.
        if kod in (401, 403, 429):
            nepusteni.append((sid, nazev, web, kod))
            continue
        if kod == 0 or kod >= 400:
            s["web_selhani"] = s.get("web_selhani", 0) + 1
            # dvoji potvrzeni: jeden vypadek neni fakt. Hlasi se presne pri
            # druhem selhani - pri tretim uz je to tataz zprava a schvalovatel
            # by kazdy mesic cetl, ze se nezmenilo nic.
            if s["web_selhani"] == 2:
                navrhy.append(navrh(sid, nazev, "web_nedostupny", T["pole_poznamka"],
                                    web,
                                    T["web_nedostupny"].format(n=s["web_selhani"], kod=kod),
                                    web, citace="HTTP %s" % kod))
            continue
        s["web_selhani"] = 0

        text = ocisti(telo, vzory_sumu)
        novy = otisk(text)
        cesta = os.path.join(stranky_dir, "%s.txt" % sid)
        stary_text = ""
        if os.path.isfile(cesta):
            with open(cesta, "r", encoding="utf-8") as f:
                stary_text = f.read()
        if s.get("otisk") and s["otisk"] != novy:
            if sid in nesledovane:
                # zmena existuje, ale vime, ze nic neznamena -> pruh C, ne fronta
                navrhy.append(navrh(sid, nazev, "text_bez_dopadu", T["pole_poznamka"],
                                    s["otisk"], novy, web,
                                    citace=nesledovane[sid]))
            else:
                zmenene.append({"id": sid, "subjekt": nazev, "url": web,
                                "stary": stary_text, "novy": text,
                                "role_fin": norm(d.get("role_financovani")),
                                "role_inv": norm(d.get("role_investor"))})
        s["otisk"] = novy
        with open(cesta, "w", encoding="utf-8") as f:
            f.write(text)

    if nepusteni:
        vypis("  Web nas nepustil dovnitr u %d subjektu (ochrana proti robotum) - "
              "NEVIME, jestli se zmenil; neni to nalez:" % len(nepusteni))
        for sid, nazev, web, kod in nepusteni:
            vypis("    #%s %s -> HTTP %s" % (sid, nazev, kod))
    if nesledovane:
        vypis("  Otisk stranky se NESLEDUJE u %d subjektu (viz otisk_ignoruj "
              "v konfiguraci):" % len(nesledovane))
        for sid, duvod in nesledovane.items():
            vypis("    #%s - %s" % (sid, duvod))

    # ---- faze 2b: mailove domeny
    # Bunka casto nese vic adres a poznamku v zavorce ("obchod@x.cz (obecny
    # office@x.cz)"). Delit retezec podle zavinace by z toho vyrobilo domenu
    # "x.cz)" a beh by kazdy mesic hlasil neexistujici domenu.
    MAIL = re.compile(r"[A-Za-z0-9._%+-]+@([A-Za-z0-9-]+(?:\.[A-Za-z0-9-]+)+)")
    zive_id = set(str(d.get("id")) for _, d in zive)
    padle = {}      # sid -> (subjekt, domena, cela bunka)
    zdrave = set()

    for radek, d in sesit.radky("kontakty"):
        sid = str(d.get("id"))
        if sid not in zive_id:
            continue
        mail = norm(d.get("email"))
        for domena in set(MAIL.findall(mail)):
            if sit.dns_ok(domena):
                zdrave.add(sid)
            else:
                padle.setdefault(sid, (norm(d.get("subjekt")), domena, mail))

    # pocitadlo se zvedne nejvyse jednou za beh a jeden subjekt, jinak by
    # subjekt se dvema kontaktnimi radky "selhal dvakrat po sobe" hned poprve
    # a dvoji potvrzeni by neplatilo
    for sid in zdrave - set(padle):
        st_subj.setdefault(sid, {})["mail_selhani"] = 0
    for sid, (subjekt, domena, mail) in padle.items():
        s = st_subj.setdefault(sid, {})
        s["mail_selhani"] = s.get("mail_selhani", 0) + 1
        if s["mail_selhani"] == 2:
            navrhy.append(navrh(sid, subjekt, "mail_domena_neexistuje",
                                T["pole_email"], mail,
                                T["mail_domena"].format(domena=domena,
                                                        n=s["mail_selhani"]),
                                "DNS", citace=T["dns_bez_zaznamu"]))

    return navrhy, zmenene


# ---------------------------------------------------------------- faze 4

def priprav_k_precteni(zmenene, cfg, stav):
    """Zapise rozdily zmenenych stranek + zadani pro model. Sam nic necte."""
    d = os.path.join(HERE, cfg["k_precteni"])
    if os.path.isdir(d):
        shutil.rmtree(d)
    os.makedirs(d)

    limit = cfg["rotace"]["max_k_precteni"]
    vybrane = zmenene[:limit]
    vynechane = zmenene[limit:]

    for z in vybrane:
        rozdil = list(difflib.unified_diff(
            z["stary"].splitlines(), z["novy"].splitlines(),
            fromfile="minule", tofile="ted", lineterm="", n=2))
        zajimave = [r for r in rozdil if r.startswith(("+", "-")) and not r.startswith(("+++", "---"))]
        role = []
        if z.get("role_fin"):
            role.append("poskytovatel financovani (list 3)")
        if z.get("role_inv"):
            role.append("investor (list 6)")
        obsah = ["# %s (#%s)" % (z["subjekt"], z["id"]),
                 "", "URL: %s" % z["url"],
                 "Role: %s" % (", ".join(role) or "neurcena"), "",
                 "Zmenenych radku: %d" % len(zajimave), "",
                 "## Rozdil proti minulemu behu", "", "```diff"]
        obsah += rozdil if rozdil else ["(bez textoveho rozdilu)"]
        obsah += ["```"]
        with open(os.path.join(d, "%s.md" % z["id"]), "w", encoding="utf-8") as f:
            f.write("\n".join(obsah))

    zadani = ZADANI_FAZE4
    if vynechane:
        zadani += ("\n\nPOZOR: zmenenych stranek bylo %d, do tehle davky jich islo %d "
                   "(strop max_k_precteni). Nezpracovane subjekty: %s. "
                   "Nejde o 'nic se nenaslo' - jde o vedome useknuti.\n"
                   % (len(zmenene), len(vybrane),
                      ", ".join("#%s" % z["id"] for z in vynechane)))
    with open(os.path.join(d, "_ZADANI.md"), "w", encoding="utf-8") as f:
        f.write(zadani)
    return len(vybrane), len(vynechane)


ZADANI_FAZE4 = """# Zadani pro fazi 4 - cteni zmenenych stranek

V teto slozce je jeden soubor na kazdy subjekt, jehoz web se od minuleho behu
zmenil. Kazdy soubor obsahuje rozdil ocisteneho textu stranky.

Tvoje uloha: rozhodnout, jestli zmena textu znamena zmenu NEKTERE Z EVIDOVANYCH
HODNOT. Vetsina zmen na webech zadnou nezname - novy clanek, jina fotka, jine
poradi odstavcu. To je v poradku a spravna odpoved je "nic".

Sleduj jen tato pole:
  typ_financovani   novy nebo zruseny typ (senior, whole loan, junior, mezzanine,
                    bridge, pref. equity, development, akvizicni, refinancovani,
                    NAV lending, financovani SPV, financovani fondu)
  ticket            minimalni nebo maximalni velikost financovani
  ltv               maximalni LTV
  kontaktni_osoba   jmenovita osoba, jeji pozice, telefon, e-mail
  transakce         dolozena nova transakce (kdo, kolik, kdy)

Kdyz ma subjekt v hlavicce souboru roli "investor (list 6)", sleduj navic:
  segment           do ceho investuje (nemovitosti, private debt, private
                    equity, venture, dluhopisy...)
  aum               objem spravovanych aktiv
  gatekeeper        pres koho se k nemu chodi - platforma, poradce, banka

U subjektu, ktery ma obe role, sleduj obe skupiny poli. Roli si nevymyslej -
je v hlavicce souboru a bere se z listu 1.

PRAVIDLO, KTERE SE NEPORUSUJE:
Ke kazdemu navrhu musis dodat DOSLOVNOU CITACI ze stranky a URL. Kdyz neumis
citovat, navrh nevznika. Necituj z rozdilu to, co v nem neni.

Nerozhoduj o tom, do ktereho pruhu navrh pujde. Pruh urcuje skript podle toho,
ktere pole se meni - ne podle toho, jak si jisty si ty.

Vystup zapis do souboru navrhy.json vedle skriptu, v tomto tvaru:

[
  {
    "id": "12",
    "subjekt": "Nazev z hlavicky souboru",
    "druh": "ticket",
    "co": "Ticket do",
    "bylo": "80 mil. Kc",
    "navrzeno": "125 mil. Kc",
    "zdroj": "https://...",
    "citace": "doslovna veta ze stranky",
    "jistota": "vysoka"
  }
]

Pak spust:  python financovani-beh.py --navrhy navrhy.json --zapis
"""


# ---------------------------------------------------------------- faze 5

def zapis(sesit, navrhy, cfg, stav, opravdu):
    """Zaradi navrhy do pruhu a zapise je. Vraci prehled."""
    scfg = cfg["sloupce"]["subjekty"]
    subjekty = {str(d.get("id")): (r, d) for r, d in sesit.radky("subjekty")}
    pocet_radku = max(1, len(subjekty))

    # co uz schvalovatel jednou zamitnul, se nenabizi podruhe
    pamet = stav.get("neopakovat", {})
    if pamet:
        pred = len(navrhy)
        navrhy = [n for n in navrhy if klic_neopakuj(n) not in pamet]
        if pred != len(navrhy):
            vypis("Preskoceno %d navrhu, ktere uz byly jednou zamitnuty."
                  % (pred - len(navrhy)))

    a, b, c = [], [], []
    for n in navrhy:
        pruh = routuj(n, cfg)
        if pruh == "B" and pod_prahem(n, cfg):
            pruh = "C"
        n["pruh"] = pruh
        {"A": a, "B": b, "C": c}[pruh].append(n)

    # pojistka: strop pruhu A
    strop = cfg["pojistky"]["strop_pruh_a_procent"]
    if len(a) * 100.0 / pocet_radku > strop:
        vypis("POJISTKA: pruh A chce zmenit %d radku z %d (vic nez %d %%). "
              "Nic se automaticky nezapise, vsechno jde ke schvaleni."
              % (len(a), pocet_radku, strop))
        for n in a:
            n["pruh"] = "B"
            n["co"] = n["co"] + " " + T["zastaveno_stropem"]
        b = a + b
        a = []

    if not opravdu:
        return {"A": a, "B": b, "C": c, "zapsano": False}

    ws = sesit.ws("subjekty")
    log = stav.setdefault("log", [])

    # pruh A: meni jen pole Stav a pripise poznamku. Nikdy nemaze radek.
    for n in a:
        if n["id"] not in subjekty:
            continue
        radek, d = subjekty[n["id"]]
        i_stav = sesit.sl("subjekty", "stav")
        i_pozn = sesit.sl("subjekty", "poznamka")
        i_over = sesit.sl("subjekty", "overeno")
        bylo = ws.cell(row=radek, column=i_stav).value
        if n["druh"] in ("insolvence", "zanik", "likvidace"):
            ws.cell(row=radek, column=i_stav).value = n["navrzeno"]
        pozn = ws.cell(row=radek, column=i_pozn).value or ""
        radek_pozn = T["automat_poznamka"].format(
            datum=DNES, co=n["co"], detail=n["citace"] or n["navrzeno"], zdroj=n["zdroj"])
        ws.cell(row=radek, column=i_pozn).value = (pozn + "\n" + radek_pozn).strip()
        if i_over:
            ws.cell(row=radek, column=i_over).value = DNES
        log.append({"datum": DNES, "id": n["id"], "pole": "stav",
                    "bylo": bylo, "nove": n["navrzeno"], "pruh": "A",
                    "druh": n["druh"], "zdroj": n["zdroj"], "radek": radek})

    # pruh B: do fronty ke schvaleni
    if b:
        zapis_navrhy(sesit, b, cfg)

    # pruh C: jen do logu
    for n in c:
        log.append({"datum": DNES, "id": n["id"], "pole": n["co"],
                    "bylo": n["bylo"], "nove": n["navrzeno"], "pruh": "C",
                    "druh": n["druh"], "zdroj": n["zdroj"]})

    return {"A": a, "B": b, "C": c, "zapsano": True}


def zapis_navrhy(sesit, navrhy, cfg):
    """Zapise pruh B do listu Navrhy zmen. Predvyplni rozhodnuti."""
    ws = sesit.ws("navrhy")
    mapa = cfg["sloupce"]["navrhy"]
    h = sesit.hlavicka("navrhy")

    # chybejici sloupce smi pribyt (pridavat lze, prejmenovat ne)
    if zajisti_sloupce(sesit, "navrhy"):
        h = sesit.hlavicka("navrhy")
    roletka(sesit, cfg)

    # ukazkove radky ze sablony pryc - sesit sam v listu 5 rika, ze se maji smazat
    i_subj = h.get(mapa["subjekt"])
    if i_subj:
        for r in range(ws.max_row, 1, -1):
            v = norm(ws.cell(row=r, column=i_subj).value)
            if v.startswith(T["ukazkovy_radek"]) or v.startswith("Ukazkov"):
                ws.delete_rows(r)

    r = ws.max_row + 1
    while r > 2 and all(ws.cell(row=r - 1, column=i).value in (None, "")
                        for i in range(1, ws.max_column + 1)):
        r -= 1

    for n in navrhy:
        hodnoty = {
            "datum": DNES,
            "id": n["id"],
            "subjekt": n["subjekt"],
            "co": n["co"],
            "bylo": n["bylo"],
            "navrzeno": n["navrzeno"],
            "zdroj": n["zdroj"],
            "schvalit": predvyplneno(n, cfg),
            "poznamka": n.get("citace", ""),
            "pruh": n["pruh"],
            "druh": n["druh"],
            "vyrizeno": "",
        }
        for pole, hodnota in hodnoty.items():
            i = h.get(mapa[pole])
            if i:
                ws.cell(row=r, column=i).value = hodnota
        r += 1


def predvyplneno(n, cfg):
    """Dokud byl sloupec Schvalit dekorace, davalo predvyplneni smysl. Ted je to
    spoustec: co je predvyplnene jako 'prijmout', to ZAPSAT.cmd opravdu zapise.
    Kdo se na frontu nepodiva, ten ji tim odsouhlasil - proto ve vychozim stavu
    zustava prazdne a schvalovatel vybira z roletky."""
    if not cfg["pojistky"].get("predvyplnovat_schvaleni"):
        return ""
    if n.get("jistota") in ("vysoka", "high"):
        return T["prijmout"]
    if n.get("jistota") in ("nizka", "low"):
        return T["zamitnout"]
    if n["druh"] in ("ico_nesedi", "novy_subjekt"):
        return ""
    return T["prijmout"]


# ---------------------------------------------------------------- odtok fronty
# Fronta, ktera se jen plni, prestane byt frontou a stane se seznamem. Tahle
# cast ji vyprazdnuje: co schvalovatel v listu 5 oznaci, to se zapise tam, kam
# to patri, nebo se natrvalo zapamatuje jako "uz nenabizet".


def klic_neopakuj(n):
    """Identita navrhu. Kdyz se hodnota na webu zmeni znovu, je to novy navrh."""
    return "%s|%s|%s|%s" % (n.get("id"), norm(n.get("druh")),
                            norm(n.get("co")), norm(n.get("navrzeno")))


def zajisti_sloupce(sesit, klic):
    """Chybejici sloupce z konfigurace dopise do hlavicky. Pridavat lze, prejmenovat ne."""
    ws = sesit.ws(klic)
    mapa = sesit.cfg["sloupce"][klic]
    h = sesit.hlavicka(klic)
    pribylo = []
    for pole in mapa:
        nazev = mapa[pole]
        if norm(nazev) not in h:
            ws.cell(row=1, column=ws.max_column + 1).value = nazev
            pribylo.append(nazev)
    if pribylo:
        sesit._hlavicky.pop(klic, None)
    return pribylo


def zajisti_list(sesit, klic):
    """Zalozi list, kdyz v sesitu jeste neni. Struktura se tim nemeni - pridava se."""
    nazev = sesit.listy[klic]
    if nazev in sesit.wb.sheetnames:
        zajisti_sloupce(sesit, klic)
        return False
    ws = sesit.wb.create_sheet(nazev)
    mapa = sesit.cfg["sloupce"][klic]
    for i, pole in enumerate(mapa, start=1):
        ws.cell(row=1, column=i).value = mapa[pole]
    sesit._hlavicky.pop(klic, None)
    return True


def roletka(sesit, cfg):
    """Sloupec Schvalit jako rozbalovaci seznam - schvalovatel nic nepise."""
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    ws = sesit.ws("navrhy")
    i = sesit.sl("navrhy", "schvalit")
    if not i:
        return False
    pismeno = get_column_letter(i)
    rozsah = "%s2:%s1000" % (pismeno, pismeno)
    # stara validace na tomtez sloupci pryc, at jich tam nesedi pet pres sebe
    for dv in list(ws.data_validations.dataValidation):
        if pismeno in str(dv.sqref or ""):
            ws.data_validations.dataValidation.remove(dv)
    dv = DataValidation(type="list",
                        formula1='"%s,%s"' % (T["prijmout"], T["zamitnout"]),
                        allow_blank=True,
                        showDropDown=False)   # OOXML naopak: False = sipka se ZOBRAZI
    dv.promptTitle = T["roletka_titulek"]
    dv.prompt = T["roletka_zprava"]
    dv.errorTitle = T["roletka_titulek"]
    dv.error = T["roletka_chyba"]
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(rozsah)
    return True


def najdi_radek(sesit, klic, sid):
    """Radek subjektu v listu podle ID. Vraci cislo radku nebo None."""
    i_id = sesit.sl(klic, "id")
    if not i_id:
        return None
    ws = sesit.ws(klic)
    for r in range(2, ws.max_row + 1):
        if norm(ws.cell(row=r, column=i_id).value) == norm(sid):
            return r
    return None


def smaz_vysvetlivku(sesit, klic):
    """Prazdny list casto nese pod hlavickou vetu pro cloveka ('sem se preklopi
    ...'). Jakmile do nej pribudou data, je z te vety falesny prvni radek."""
    ws = sesit.ws(klic)
    if ws.max_row < 2:
        return
    for r in range(2, min(ws.max_row, 5) + 1):
        bunky = [ws.cell(row=r, column=i).value for i in range(1, ws.max_column + 1)]
        vyplnene = [i for i, v in enumerate(bunky) if not prazdne(v)]
        if len(vyplnene) == 1 and vyplnene[0] == 0 and len(norm(bunky[0])) > 60:
            ws.delete_rows(r)
            return


def zaloz_radek(sesit, klic, sid, subjekt):
    """Podrizeny list nema radek pro kazdy subjekt - kdyz chybi, zalozi se."""
    smaz_vysvetlivku(sesit, klic)
    ws = sesit.ws(klic)
    r = ws.max_row + 1
    while r > 2 and all(ws.cell(row=r - 1, column=i).value in (None, "")
                        for i in range(1, ws.max_column + 1)):
        r -= 1
    i_id = sesit.sl(klic, "id")
    if i_id:
        ws.cell(row=r, column=i_id).value = sid
    for pole in ("subjekt", "nazev"):
        i = sesit.sl(klic, pole)
        if i:
            ws.cell(row=r, column=i).value = subjekt
            break
    return r


def zapis_zdroj(sesit, sid, subjekt, k_cemu, citace, url):
    """Kazda zapsana hodnota musi mit v listu 4 doslovnou citaci a URL."""
    if not citace or not url:
        return
    mapa = sesit.cfg["sloupce"].get("zdroje")
    if not mapa or "zdroje" not in sesit.listy:
        return
    if sesit.listy["zdroje"] not in sesit.wb.sheetnames:
        return
    ws = sesit.ws("zdroje")
    h = sesit.hlavicka("zdroje")
    r = ws.max_row + 1
    hodnoty = {"id": sid, "subjekt": subjekt, "k_cemu": k_cemu,
               "citace": citace, "url": url, "overeno": DNES}
    for pole in hodnoty:
        nazev = mapa.get(pole)
        if not nazev:
            continue
        i = h.get(norm(nazev))
        if i:
            ws.cell(row=r, column=i).value = hodnoty[pole]


def pripis_poznamku(sesit, radek, text):
    i = sesit.sl("subjekty", "poznamka")
    if not i:
        return
    ws = sesit.ws("subjekty")
    stara = ws.cell(row=radek, column=i).value or ""
    ws.cell(row=radek, column=i).value = (stara + "\n" + text).strip()


def aplikuj_jeden(sesit, cfg, d, stav, opravdu):
    """Zapise jeden schvaleny navrh. Vraci (povedlo se, kam to slo)."""
    druh = norm(d.get("druh"))
    pravidlo = cfg["aplikace"].get(druh)
    if not isinstance(pravidlo, dict):
        return False, "druh '%s' nema pravidlo v sekci aplikace" % (druh or "?")

    klic = pravidlo["list"]
    sloupec = pravidlo["sloupec"]
    sid = norm(d.get("id"))
    subjekt = norm(d.get("subjekt"))
    co = norm(d.get("co"))
    hodnota = norm(d.get("navrzeno"))
    citace = norm(d.get("poznamka"))
    zdroj = norm(d.get("zdroj"))
    log = stav.setdefault("log", [])
    popis = T["aplikovano_pozn"].format(datum=DNES, co=co,
                                        detail=citace or hodnota, zdroj=zdroj)

    # --- novy subjekt: novy radek v listu 1
    if sloupec == "novy_radek":
        if not opravdu:
            return True, "%s (novy radek)" % sesit.listy["subjekty"]
        ws = sesit.ws("subjekty")
        i_id = sesit.sl("subjekty", "id")
        nejvyssi = 0
        for r in range(2, ws.max_row + 1):
            v = cislo(ws.cell(row=r, column=i_id).value)
            if v and v > nejvyssi:
                nejvyssi = v
        nove_id = int(nejvyssi) + 1
        r = zaloz_radek(sesit, "subjekty", nove_id, subjekt)
        for pole, val in (("ico", sid), ("stav", cfg["stavy"]["aktivni"]), ("overeno", DNES)):
            i = sesit.sl("subjekty", pole)
            if i:
                ws.cell(row=r, column=i).value = val
        pripis_poznamku(sesit, r, popis)
        zapis_zdroj(sesit, nove_id, subjekt, co, citace, zdroj)
        log.append({"datum": DNES, "id": nove_id, "pole": "novy radek",
                    "bylo": "", "nove": subjekt, "pruh": "B", "druh": druh,
                    "zdroj": zdroj, "list": "subjekty", "radek": r,
                    "nevratne": True})
        return True, "%s (nove ID %d)" % (sesit.listy["subjekty"], nove_id)

    # --- investorska strana. Listy 1 a 2 jsou spolecny registr, list 3 nese
    # roli poskytovatele financovani a list 6 roli investora - jeden subjekt
    # muze mit obe. Zarazeni investora tedy neni novy zaznam vedle databaze,
    # ale role na radku v listu 1 plus radek v listu 6.
    if sloupec in ("role_investor", "novy_investor"):
        ws = sesit.ws("subjekty")
        if sloupec == "novy_investor":
            if not opravdu:
                return True, "%s (novy radek) + %s" % (sesit.listy["subjekty"],
                                                       sesit.listy["investor"])
            i_id = sesit.sl("subjekty", "id")
            nejvyssi = 0
            for r in range(2, ws.max_row + 1):
                v = cislo(ws.cell(row=r, column=i_id).value)
                if v and v > nejvyssi:
                    nejvyssi = v
            radek = zaloz_radek(sesit, "subjekty", int(nejvyssi) + 1, subjekt)
            sid_v_listu = int(nejvyssi) + 1
            for pole, val in (("ico", sid), ("stav", cfg["stavy"]["aktivni"]),
                              ("overeno", DNES)):
                i = sesit.sl("subjekty", pole)
                if i:
                    ws.cell(row=radek, column=i).value = val
            bylo_stav = ""
        else:
            radek = najdi_radek(sesit, "subjekty", sid)
            if not radek:
                return False, "subjekt #%s v listu 1 neni" % sid
            sid_v_listu = sid
            if not opravdu:
                return True, "%s / Role: investor + %s" % (sesit.listy["subjekty"],
                                                           sesit.listy["investor"])
            # VYRAZEN znamenalo "nepujcuje z vlastni bilance". Pro roli
            # investora je to jina otazka, tak se stav vraci na aktivni -
            # jinak by ho mesicni beh preskakoval a nehlidal.
            i_stav = sesit.sl("subjekty", "stav")
            bylo_stav = ws.cell(row=radek, column=i_stav).value
            ws.cell(row=radek, column=i_stav).value = cfg["stavy"]["aktivni"]
            i_over = sesit.sl("subjekty", "overeno")
            if i_over:
                ws.cell(row=radek, column=i_over).value = DNES

        i_role = sesit.sl("subjekty", "role_investor")
        if i_role:
            ws.cell(row=radek, column=i_role).value = T["ano"]
        pripis_poznamku(sesit, radek, T["role_investor_pozn"].format(
            datum=DNES, detail=(citace or hodnota), zdroj=zdroj))
        # radek v listu 6 - jen zalozit, hodnoty (segment, AUM) chodi zvlast
        if not najdi_radek(sesit, "investor", sid_v_listu):
            zaloz_radek(sesit, "investor", sid_v_listu, subjekt)
        zapis_zdroj(sesit, sid_v_listu, subjekt, co, citace, zdroj)
        log.append({"datum": DNES, "id": sid_v_listu, "pole": "Role: investor",
                    "bylo": bylo_stav, "nove": T["ano"], "pruh": "B", "druh": druh,
                    "zdroj": zdroj, "list": "subjekty", "radek": radek,
                    "nevratne": True})
        return True, "%s / Role: investor + %s" % (sesit.listy["subjekty"],
                                                   sesit.listy["investor"])

    # --- jen poznamka: hodnota nema cilove pole, ale stopa zustat musi
    if sloupec == "jen_poznamka":
        radek = najdi_radek(sesit, "subjekty", sid)
        if not radek:
            return False, "subjekt #%s v listu 1 neni" % sid
        if opravdu:
            pripis_poznamku(sesit, radek, popis)
            zapis_zdroj(sesit, sid, subjekt, co, citace, zdroj)
            log.append({"datum": DNES, "id": sid, "pole": T["pole_poznamka"],
                        "bylo": "", "nove": popis, "pruh": "B", "druh": druh,
                        "zdroj": zdroj, "list": "subjekty", "radek": radek,
                        "nevratne": True})
        return True, "%s / %s" % (sesit.listy["subjekty"], T["pole_poznamka"])

    # --- bezne pole: nazev sloupce bud z konfigurace, nebo ho nese sam navrh
    nazev_sl = co if sloupec == "podle_co" else sloupec
    h = sesit.hlavicka(klic)
    i = h.get(norm(nazev_sl))
    if not i:
        return False, "sloupec '%s' v listu '%s' neexistuje" % (nazev_sl, sesit.listy[klic])

    radek = najdi_radek(sesit, klic, sid)
    if not radek:
        if not opravdu:
            return True, "%s (novy radek) / %s" % (sesit.listy[klic], nazev_sl)
        radek = zaloz_radek(sesit, klic, sid, subjekt)
    if not opravdu:
        return True, "%s / %s" % (sesit.listy[klic], nazev_sl)

    ws = sesit.ws(klic)
    bylo = ws.cell(row=radek, column=i).value
    ws.cell(row=radek, column=i).value = hodnota
    i_over = sesit.sl(klic, "overeno")
    if i_over:
        ws.cell(row=radek, column=i_over).value = DNES
    zapis_zdroj(sesit, sid, subjekt, nazev_sl, citace, zdroj)
    log.append({"datum": DNES, "id": sid, "pole": nazev_sl,
                "bylo": bylo, "nove": hodnota, "pruh": "B", "druh": druh,
                "zdroj": zdroj, "list": klic, "radek": radek, "sloupec": nazev_sl})
    return True, "%s / %s" % (sesit.listy[klic], nazev_sl)


def zapis_zamitnute(sesit, polozky, kde):
    """Zamitnuti patri do sesitu, ne jen do JSONu. Za rok se nekdo zepta proc."""
    if not polozky:
        return 0
    zajisti_list(sesit, "zamitnuto")
    ws = sesit.ws("zamitnuto")
    mapa = sesit.cfg["sloupce"]["zamitnuto"]
    h = sesit.hlavicka("zamitnuto")
    if ws.max_row == 1:
        ws.cell(row=1, column=len(mapa) + 2).value = T["zamitnuto_nadpis"]
    # Klic je ICO, ale jen kdyz opravdu ICO je. Zahranicni firmy a nedohledane
    # subjekty maji v tom sloupci pomlcku - kdyby se dedupovalo podle ni,
    # zapsal by se z nich jen prvni a zbytek by zmizel beze slova.
    def klic7(ico, nazev):
        i = norm(ico)
        return i if re.fullmatch(r"\d{6,10}", i or "") else "nazev:%s" % norm(nazev).lower()

    zname = set()
    i_ico = h.get(norm(mapa["ico"]))
    i_naz = h.get(norm(mapa["nazev"]))
    if i_ico:
        for r in range(2, ws.max_row + 1):
            zname.add(klic7(ws.cell(row=r, column=i_ico).value,
                            ws.cell(row=r, column=i_naz).value if i_naz else ""))
    r = ws.max_row + 1
    pocet = 0
    for p in polozky:
        if klic7(p.get("ico"), p.get("nazev")) in zname:
            continue
        hodnoty = {"datum": p.get("datum") or DNES, "ico": p.get("ico"),
                   "nazev": p.get("nazev"), "duvod": p.get("duvod"),
                   "citace": p.get("citace"), "zdroj": p.get("zdroj"), "kde": kde}
        for pole in hodnoty:
            i = h.get(norm(mapa[pole]))
            if i:
                ws.cell(row=r, column=i).value = hodnoty[pole]
        zname.add(klic7(p.get("ico"), p.get("nazev")))
        r += 1
        pocet += 1
    return pocet


def aplikuj(sesit, cfg, stav, opravdu):
    """Projde list 5 a vyridi vsechno, u ceho schvalovatel rozhodl."""
    zajisti_sloupce(sesit, "navrhy")
    ws5 = sesit.ws("navrhy")
    mapa5 = cfg["sloupce"]["navrhy"]
    h5 = sesit.hlavicka("navrhy")
    i_vyriz = h5.get(norm(mapa5["vyrizeno"]))

    pamet = stav.setdefault("neopakovat", {})
    prijato, zamitnuto, nezname = [], [], []
    ceka = 0
    do_listu7 = []

    ano = norm(T["prijmout"]).lower()
    ne = norm(T["zamitnout"]).lower()

    for r, d in sesit.radky("navrhy"):
        if not prazdne(d.get("vyrizeno")):
            continue
        subj = norm(d.get("subjekt"))
        # pod tabulkou byva poznamka pro cloveka - neni to navrh, jen text
        if prazdne(d.get("id")) and prazdne(subj):
            continue
        if subj.startswith(T["ukazkovy_radek"]) or subj.startswith("Ukazkov"):
            continue
        rozhodnuti = norm(d.get("schvalit")).lower()
        if rozhodnuti not in (ano, ne):
            if rozhodnuti:
                nezname.append((r, "ve sloupci Schvalit stoji '%s'" % rozhodnuti))
            else:
                ceka += 1
            continue

        if rozhodnuti == ano:
            ok, kam = aplikuj_jeden(sesit, cfg, d, stav, opravdu)
            if not ok:
                nezname.append((r, kam))
                continue
            prijato.append((d, kam))
        else:
            pamet[klic_neopakuj(d)] = {
                "datum": DNES, "id": norm(d.get("id")), "subjekt": subj,
                "co": norm(d.get("co")), "navrzeno": norm(d.get("navrzeno"))}
            zamitnuto.append(d)
            if norm(d.get("druh")) == "novy_subjekt":
                do_listu7.append({"datum": DNES, "ico": norm(d.get("id")),
                                  "nazev": subj, "duvod": norm(d.get("poznamka")),
                                  "citace": norm(d.get("poznamka")),
                                  "zdroj": norm(d.get("zdroj"))})

        if opravdu and i_vyriz:
            ws5.cell(row=r, column=i_vyriz).value = DNES

    do7 = 0
    if opravdu and do_listu7:
        do7 = zapis_zamitnute(sesit, do_listu7, T["zamitnuto_kde_5"])

    return {"prijato": prijato, "zamitnuto": zamitnuto, "ceka": ceka,
            "nezname": nezname, "do_listu7": do7, "zapsano": opravdu}


def prehled_aplikace(v, sesit):
    r = ["", "=" * 64, T["nadpis_aplikace"], "=" * 64]
    r.append("  prijato a zapsano: %d" % len(v["prijato"]))
    for d, kam in v["prijato"][:40]:
        r.append("    #%s %s -> %s: %s" % (d.get("id"), norm(d.get("subjekt"))[:34],
                                           kam, norm(d.get("navrzeno"))[:40]))
    r.append("  zamitnuto (uz se nenabidne): %d" % len(v["zamitnuto"]))
    for d in v["zamitnuto"][:40]:
        r.append("    #%s %s - %s" % (d.get("id"), norm(d.get("subjekt"))[:34],
                                      norm(d.get("co"))[:30]))
    if v["do_listu7"]:
        r.append("  do listu '%s' pribylo: %d"
                 % (sesit.listy["zamitnuto"], v["do_listu7"]))
    r.append("  ceka na rozhodnuti: %d" % v["ceka"])
    if v["nezname"]:
        r.append("  NEVYRIZENO - skript nevi kam s tim: %d" % len(v["nezname"]))
        for radek, duvod in v["nezname"][:20]:
            r.append("    radek %s: %s" % (radek, duvod))
        r.append("  Tyhle radky zustavaji ve fronte. Bud je to preklep ve sloupci")
        r.append("  Schvalit, nebo chybi pravidlo v sekci 'aplikace' v konfiguraci.")
    if not v["zapsano"]:
        r.append("")
        r.append("  (nanecisto - nic se nezapsalo; pro skutecny zapis pridej --zapis)")
    return "\n".join(r)

# ---------------------------------------------------------------- vraceni

def vrat(sesit, stav, datum, opravdu):
    log = stav.get("log", [])
    k_vraceni = [z for z in log if z["datum"] == datum and z["pruh"] in ("A", "B")]
    if not k_vraceni:
        vypis("Z behu %s neni co vracet (zadna zapsana zmena)." % datum)
        return
    vracene, nevratne = [], []
    for z in k_vraceni:
        # novy radek a pripsana poznamka se nevraci - mazat radek z databaze
        # kvuli kroku zpet je horsi nez ho tam nechat a rict o nem nahlas
        if z.get("nevratne"):
            nevratne.append(z)
            continue
        klic = z.get("list", "subjekty")
        ws = sesit.ws(klic)
        if z.get("sloupec"):
            i = sesit.hlavicka(klic).get(norm(z["sloupec"]))
        else:
            i = sesit.sl("subjekty", "stav")
        if not i or not z.get("radek"):
            nevratne.append(z)
            continue
        vypis("  #%s %s: '%s' -> zpet '%s'"
              % (z["id"], z.get("pole", ""), z["nove"], z["bylo"]))
        if opravdu:
            ws.cell(row=z["radek"], column=i).value = z["bylo"]
        vracene.append(z)
    if nevratne:
        vypis("  NEVRACI SE %d zaznamu (novy radek nebo pripsana poznamka):" % len(nevratne))
        for z in nevratne:
            vypis("    #%s %s: %s" % (z["id"], z.get("pole", ""), str(z["nove"])[:60]))
        vypis("  Tohle musis vzit zpatky rucne - zaloha sesitu je ve slozce zalohy\\.")
    if opravdu:
        vracene_id = [id(z) for z in vracene]
        stav["log"] = [z for z in log if id(z) not in vracene_id]
        vypis("Vraceno %d zmen." % len(vracene))
    else:
        vypis("(nanecisto - pro skutecne vraceni pridej --zapis)")


# ---------------------------------------------------------------- prehled

def prehled(vysledek, zmenene, k_precteni, cfg, stav):
    a, b, c = vysledek["A"], vysledek["B"], vysledek["C"]
    radky = []
    radky.append("")
    radky.append("=" * 62)
    radky.append("MESICNI BEH %s" % DNES)
    radky.append("=" * 62)
    radky.append("Pruh A (aplikovano automaticky): %d" % len(a))
    for n in a:
        radky.append("   #%s %s | %s: %s -> %s" % (n["id"], n["subjekt"][:40],
                                                   n["co"], n["bylo"], n["navrzeno"]))
    radky.append("Pruh B (ceka na schvaleni):      %d" % len(b))
    for n in b:
        radky.append("   #%s %s | %s: %s -> %s" % (n["id"], n["subjekt"][:40],
                                                   n["co"], n["bylo"], n["navrzeno"]))
    radky.append("Pruh C (jen do logu):            %d" % len(c))
    radky.append("Zmenenych stranek:               %d" % len(zmenene))
    if k_precteni:
        radky.append("Pripraveno k precteni modelem:   %d (slozka %s)"
                     % (k_precteni[0], cfg["k_precteni"]))
        if k_precteni[1]:
            radky.append("   POZOR: %d zmenenych stranek se do davky neveslo "
                         "(strop max_k_precteni) - neni to 'nic se nenaslo'."
                         % k_precteni[1])
    radky.append("")

    # mereni vlastniho provozu - fronta a pomer zamitnuti ridi serizeni prahu
    if len(b) > 25:
        radky.append("Fronta ke schvaleni je nad 25 polozek -> pritvrdit prahy "
                     "v konfiguraci (sekce prahy).")
    if not (a or b or c):
        radky.append("Zkontrolovano bez nalezu. Nic ke schvaleni.")
        radky.append("(Tahle zprava chodi i kdyz se nic nezmenilo - ticho je "
                     "dvojznacne, nepozna se klid od spadleho behu.)")
    if not vysledek["zapsano"]:
        radky.append("NANECISTO - nic nebylo zapsano. Ostry beh: --zapis")
    return "\n".join(radky)


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description="Mesicni beh nad databazi poskytovatelu financovani")
    ap.add_argument("--zapis", action="store_true", help="opravdu zapsat (jinak jen nanecisto)")
    ap.add_argument("--jen-registry", action="store_true", help="jen faze 1 (ARES + ISIR)")
    ap.add_argument("--limit", type=int, help="jen prvnich N subjektu (test)")
    ap.add_argument("--navrhy", help="soubor navrhy.json z faze 4")
    ap.add_argument("--vrat", help="vrati zapsane zmeny z behu daneho data (YYYY-MM-DD)")
    ap.add_argument("--master", help="jina cesta k sesitu (test, kdyz je disk O: pryc)")
    ap.add_argument("--aplikovat", action="store_true",
                    help="vyridi list 5: co je schvalene, zapise; co zamitnute, zapamatuje")
    args = ap.parse_args()

    cfg = load_json(CONFIG, None)
    if cfg is None:
        raise SystemExit("Chybi %s" % CONFIG)
    T.update(cfg["texty"])
    stav = load_json(STATE, {"behy": [], "subjekty": {}, "log": []})

    if args.master:
        cfg["master"] = args.master

    sesit = Sesit(cfg)

    if args.vrat:
        vrat(sesit, stav, args.vrat, args.zapis)
        if args.zapis:
            sesit.uloz(os.path.join(HERE, cfg["zalohy"]))
            save_json(STATE, stav)
        return

    if args.aplikovat:
        zajisti_list(sesit, "zamitnuto")
        roletka(sesit, cfg)
        vysledek = aplikuj(sesit, cfg, stav, args.zapis)
        vypis(prehled_aplikace(vysledek, sesit))
        if args.zapis:
            sesit.uloz(os.path.join(HERE, cfg["zalohy"]))
            save_json(STATE, stav)
        return

    if args.navrhy:
        vstup = load_json(args.navrhy, [])
        cizi = []
        for n in vstup:
            # pravidlo: bez doslovne citace a URL navrh nevznika
            if not n.get("citace") or not n.get("zdroj"):
                cizi.append(n)
                continue
            n.setdefault("bylo", "")
            n.setdefault("co", n.get("druh", ""))
        vstup = [n for n in vstup if n not in cizi]
        if cizi:
            vypis("Zahozeno %d navrhu bez doslovne citace nebo bez URL." % len(cizi))
        vysledek = zapis(sesit, vstup, cfg, stav, args.zapis)
        vypis(prehled(vysledek, [], None, cfg, stav))
        if args.zapis:
            sesit.uloz(os.path.join(HERE, cfg["zalohy"]))
            save_json(STATE, stav)
        return

    sit = Sit(cfg)
    navrhy, zmenene = projdi(sesit, sit, stav, cfg,
                             limit=args.limit, jen_registry=args.jen_registry)

    k_precteni = None
    if zmenene:
        k_precteni = priprav_k_precteni(zmenene, cfg, stav)

    vysledek = zapis(sesit, navrhy, cfg, stav, args.zapis)
    vypis(prehled(vysledek, zmenene, k_precteni, cfg, stav))

    stav["posledni_beh"] = DNES
    stav.setdefault("behy", []).append({
        "datum": DNES,
        "nalezeno": len(navrhy),
        "pruh_a": len(vysledek["A"]),
        "pruh_b": len(vysledek["B"]),
        "pruh_c": len(vysledek["C"]),
        "zmenenych_stranek": len(zmenene),
        "zapsano": vysledek["zapsano"],
    })
    if args.zapis:
        sesit.uloz(os.path.join(HERE, cfg["zalohy"]))
    save_json(STATE, stav)


if __name__ == "__main__":
    main()
